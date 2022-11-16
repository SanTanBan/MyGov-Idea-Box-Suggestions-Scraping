#https://www.youtube.com/watch?v=lkPqPqJCElE


# Tutorial 1 telling to download ChromeDriver
# https://www.youtube.com/watch?v=Xjv1sY630Uc&list=PLzMcBGfZo4-n40rB1XaJ0ak1bemvlqumQ&index=2 


# Tutorial 2: Accessing Elements in a WebPage:~ ID, Class, Name, Tag
# https://www.youtube.com/watch?v=b5jt2bhSeXs&t=173s
# If I search using the Class name, Selenium would return the first such result

from os import system
#system("pip install selenium")

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl

PATH = "C:\Program Files (x86)\chromedriver.exe"
browser = webdriver.Chrome(PATH)
#from webdriver_manager.chrome import ChromeDriverManager
#browser = webdriver.Chrome(ChromeDriverManager().install())


#browser.get("https://www.blupower.in/")
#browser.get("https://techwithtim.net/")
browser.get("https://www.mygov.in/group-issue/mygov-idea-box/")

a=7
time.sleep(a) # Give some website loading time
print("\n")
print(browser.title)
print("\n")
#print(browser.page_source)
#print("\n")

##search = browser.find_element_by_name("Awards")
#search = browser.find_element("name","s")
#search.send_keys("test") # Whatever you want to type within a search box
#search.send_keys(Keys.RETURN)

#button = browser.find_element_by_link_text('View More')

wb_individual = openpyxl.Workbook()
sheet_individual = wb_individual.active
cell = sheet_individual.cell(row = 1, column = 1)
cell.value = "User Name"
cell = sheet_individual.cell(row = 1, column = 2)
cell.value = "Suggestion"

i=1
while i:
    # How to use find_element
    # https://stackoverflow.com/questions/69875125/find-element-by-commands-are-deprecated-in-selenium
    try:
        button = browser.find_element(By.LINK_TEXT,"View More")
        button.click()
        time.sleep(a) # Give some website loading time
        b=1

        #comment_content = browser.find_elements(By.ID,"comment_main_content_*")
        all_comment_contents = browser.find_elements(By.CLASS_NAME,"comment_body")
        all_comment_usernames = browser.find_elements(By.CLASS_NAME,"username")

        for row_number,individual_comment_content in enumerate(all_comment_contents):  
            cell = sheet_individual.cell(row = row_number + 3, column = 2)
            cell.value = individual_comment_content.text

        for row_number,individual_username in enumerate(all_comment_usernames):  
            cell = sheet_individual.cell(row = row_number + 3, column = 1)
            cell.value = individual_username.text

        wb_individual.save("Nov14.xlsx")

        i+=1

        if i%5==0:
            system('copy "D:\\Scraping MyGov\\Nov14.xlsx" "D:\\Scraping MyGov\\Nov14_BackUp.xlsx"')

        elif i%11==0:
            system('copy "D:\\Scraping MyGov\\Nov14_BackUp.xlsx" "D:\\Scraping MyGov\\Nov14_Double_BackUp.xlsx"')

        elif i%23==0:
            system('copy "D:\\Scraping MyGov\\Nov14_Double_BackUp.xlsx" "D:\\Scraping MyGov\\Nov14_Triple_BackUp.xlsx" ')

        elif i%47==0:
            system('copy "D:\\Scraping MyGov\\Nov14_Triple_BackUp.xlsx" "D:\\Scraping MyGov\\Nov14_Fourth_BackUp.xlsx" ')


    except:
        print("Inside Exception in Line 93")
        a+=2
        b=0
    if b:
        a=7


#browser.close()
#browser.quit()
